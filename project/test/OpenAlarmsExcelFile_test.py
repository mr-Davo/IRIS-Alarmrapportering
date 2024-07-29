import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def CreateExcelFile(csv_file):
    ExcelFileName= "OpenAlarms.xlsx"
    df1 = pd.read_csv(csv_file,sep=";")
    file_name = csv_file.split(".")[0]
    with pd.ExcelWriter(ExcelFileName, engine='xlsxwriter') as writer:
        df1.to_excel(writer, sheet_name=file_name, index=False)
    Format(ExcelFileName)
    return ExcelFileName
    
def Format(ExcelFile):
    
    # Open the existing workbook
    workbook = openpyxl.load_workbook(ExcelFile)

    # Get the first sheet
    sheet = workbook.worksheets[0]
    workbook.active = 0

    # Add filters to the first 5 columns
    sheet.auto_filter.ref = sheet.dimensions
    
    # Set the background color of the first row to light grey
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = grey_fill
        # Set the width for every column on sheet 2
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 10
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 85
        sheet.column_dimensions["F"].width = 20
        sheet.column_dimensions["G"].width = 20
        sheet.column_dimensions["H"].width = 20
    # Block the first row
    sheet.freeze_panes = 'A2'
    # Save the modified workbook
    workbook.save(ExcelFile)