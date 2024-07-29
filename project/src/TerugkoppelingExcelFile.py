import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def CreateExcelFile(csv_files):
    ExcelFileName= "output.xlsx"
    df1 = pd.read_csv(csv_files[0],sep=";")
    df2 = pd.read_csv(csv_files[1],sep=";")
    file_name1 = csv_files[0].split(".")[0]
    file_name2 = csv_files[1].split(".")[0]
    with pd.ExcelWriter(ExcelFileName, engine='xlsxwriter') as writer:
        df1.to_excel(writer, sheet_name=file_name1, index=False)
        df2.to_excel(writer, sheet_name=file_name2, index=False)
    Format(ExcelFileName)
    return ExcelFileName
    
def Format(ExcelFile):
    
    # Open the existing workbook
    workbook = openpyxl.load_workbook(ExcelFile)
    for i in range(2):
        # Get the first sheet
        sheet = workbook.worksheets[i]

        # Make the first sheet the active worksheet
        workbook.active = i

        # Add filters to the first 5 columns
        sheet.auto_filter.ref = sheet.dimensions
        
        # Set the background color of the first row to light grey
        grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in sheet[1]:
            cell.fill = grey_fill
        if i == 0:
            # Set the width for every column on sheet 1
            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 100
            sheet.column_dimensions["C"].width = 20
            sheet.column_dimensions["D"].width = 20
            sheet.column_dimensions["E"].width = 12.5
            sheet.column_dimensions["F"].width = 20
            sheet.column_dimensions["G"].width = 20
            sheet.column_dimensions["H"].width = 12.5
        if i == 1:
            # Set the width for every column on sheet 2
            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 10
            sheet.column_dimensions["C"].width = 15
            sheet.column_dimensions["D"].width = 15
            sheet.column_dimensions["E"].width = 85
            sheet.column_dimensions["F"].width = 12.5
            sheet.column_dimensions["G"].width = 15
            sheet.column_dimensions["H"].width = 15
            sheet.column_dimensions["I"].width = 15
        # Block the first row
        sheet.freeze_panes = 'A2'
    # Save the modified workbook
    workbook.save(ExcelFile)