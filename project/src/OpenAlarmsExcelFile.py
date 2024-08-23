import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime


def CreateExcelFile(csv_file):
    ExcelFileName= "OpenAlarms.xlsx"
    df1 = pd.read_csv(csv_file,sep=";")
    row = df1.shape[0]
    df1['Aanvangstijd'] = pd.to_datetime(df1['Aanvangstijd'], format='%d/%m/%Y %H:%M:%S')
    df1['Tijd bevestigd'] = pd.to_datetime(df1['Tijd bevestigd'], format='%d/%m/%Y %H:%M:%S')
    file_name = csv_file.split(".")[0]    
    with pd.ExcelWriter(ExcelFileName, engine='xlsxwriter') as writer:
        df1.to_excel(writer, sheet_name=file_name, index=False)
    Format(ExcelFileName,row)
    return ExcelFileName
    
def Format(ExcelFile,rows):
    
    # Open the existing workbook
    workbook = openpyxl.load_workbook(ExcelFile)

    # Get the first sheet
    sheet = workbook.worksheets[0]
    workbook.active = 0

    sheet.insert_rows(1)
    sheet.merge_cells('A1:H1')
    uur = datetime.now().replace(microsecond=0, second=0, minute=0) 
    sheet['A1'] = f'Openstaande alarmen op tijdstip: {uur}'
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # Add filters to the first 5 columns
    print(f"A2:H{rows+2}")
    sheet.auto_filter.ref = f"A2:H{rows+2}"
    
    # Set the background color of the first row to light grey
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for row in range(1, 3):  # Rows 1 and 2
        for col in range(1,9):
            cell = sheet.cell(row=row, column=col)
            cell.fill = grey_fill

    for cell in sheet[1]:
        # Set the width for every column on sheet 2
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 10
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 85
        sheet.column_dimensions["F"].width = 20
        sheet.column_dimensions["G"].width = 20
        sheet.column_dimensions["H"].width = 20
    # Block the first row
    sheet.freeze_panes = 'A2'
    sheet.freeze_panes = 'A3'
    # Save the modified workbook
    workbook.save(ExcelFile)