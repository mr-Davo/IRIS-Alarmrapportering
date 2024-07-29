import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import multiprocessing as mp
import psutil



def get_available_memory():
    """Get available system memory in bytes."""
    return psutil.virtual_memory().available

def estimate_row_size(csv_file, n_rows=1000):
    """Estimate the size of one row by reading a small sample."""
    sample_df = pd.read_csv(csv_file, sep=';', nrows=n_rows)
    sample_memory = sample_df.memory_usage(deep=True).sum()
    return sample_memory / n_rows

def calculate_chunk_size(csv_file, memory_fraction=0.5):
    """Calculate chunk size based on available memory and estimated row size."""
    available_memory = get_available_memory() * memory_fraction
    row_size = estimate_row_size(csv_file)
    return int(available_memory / row_size)

def Process_csv_file_in_chunks(csv_file):
    chunk_size = calculate_chunk_size(csv_file, memory_fraction=0.5)
    """Read the CSV file in chunks and concatenate them."""
    chunks = pd.read_csv(csv_file, sep=';', chunksize=chunk_size)
    
    df_list = []
    for chunk in chunks:
        df_list.append(chunk)
    df = pd.concat(df_list, ignore_index=True)

def CreateExcelFile1(csv_files):
    ExcelFileName = "output.xlsx"
    try:
        # Create a pool of processes
        with mp.Pool(processes=2) as pool:
        # Map the function to the list of file paths
           results = pool.map(Process_csv_file_in_chunks, [csv_files[0], csv_files[1]])
    except Exception as e:
        print(e)
    df1 = results[0]
    df2 = results[1]
    file_name1 = csv_files[0].split(".")[0]
    file_name2 = csv_files[1].split(".")[0]
    with pd.ExcelWriter(ExcelFileName, engine='xlsxwriter') as writer:
        df1.to_excel(writer, sheet_name=file_name1, index=False)
        df2.to_excel(writer, sheet_name=file_name2, index=False)
    Format(ExcelFileName)
    return ExcelFileName


def Format(ExcelFile):
    
    # Open the existing workbook
    workbook = openpyxl.load_workbook(ExcelFile)
    for i in range(2):
        # Get the first sheet
        sheet = workbook.worksheets[i]

        # Make the first sheet the active worksheet
        workbook.active = i

        # Add filters to the first 5 columns
        sheet.auto_filter.ref = sheet.dimensions
        
        # Set the background color of the first row to light grey
        grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in sheet[1]:
            cell.fill = grey_fill
        if i == 0:
            # Set the width for every column on sheet 1
            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 100
            sheet.column_dimensions["C"].width = 15
            sheet.column_dimensions["D"].width = 20
            sheet.column_dimensions["E"].width = 20
            sheet.column_dimensions["F"].width = 12.5
            sheet.column_dimensions["G"].width = 20
        if i == 1:
            # Set the width for every column on sheet 2
            sheet.column_dimensions["A"].width = 15
            sheet.column_dimensions["B"].width = 10
            sheet.column_dimensions["C"].width = 15
            sheet.column_dimensions["D"].width = 15
            sheet.column_dimensions["E"].width = 85
            sheet.column_dimensions["F"].width = 12.5
            sheet.column_dimensions["G"].width = 15
            sheet.column_dimensions["H"].width = 15
            sheet.column_dimensions["I"].width = 15
        # Block the first row
        sheet.freeze_panes = 'A2'
    # Save the modified workbook
    workbook.save(ExcelFile)